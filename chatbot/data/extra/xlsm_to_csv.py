import pandas as pd
import os
import sys
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection


def xlsm_to_xlsx(xlsm_file_path, custom_folder_name=None, output_format="xlsx"):
    """
    Read an XLSM file and create a folder with the file name.
    Convert each sheet to XLSX/CSV files within that folder, preserving images.

    Args:
        xlsm_file_path (str): Path to the XLSM file
        custom_folder_name (str, optional): Custom name for the output folder
        output_format (str): 'xlsx' or 'csv' - format for output files
    """
    try:
        # Check if file exists
        if not os.path.exists(xlsm_file_path):
            print(f"Error: File '{xlsm_file_path}' not found.")
            return False

        # Get file name without extension or use custom name
        if custom_folder_name:
            file_name = custom_folder_name
        else:
            file_name = Path(xlsm_file_path).stem

        # Create folder with file name
        folder_path = os.path.join(os.path.dirname(xlsm_file_path), file_name)

        # Create folder if it doesn't exist
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
            print(f"Created folder: {folder_path}")
        else:
            print(f"Folder already exists: {folder_path}")

        # Load the original workbook to preserve images
        original_wb = openpyxl.load_workbook(xlsm_file_path, data_only=True)

        print(f"Processing file: {xlsm_file_path}")
        print(f"Found {len(original_wb.sheetnames)} sheets: {original_wb.sheetnames}")
        print(f"Output format: {output_format.upper()}")

        # Process each sheet
        for sheet_name in original_wb.sheetnames:
            try:
                if output_format.lower() == "xlsx":
                    # Create individual XLSX file for each sheet with images and formatting
                    sheet_wb = Workbook()
                    sheet_ws = sheet_wb.active
                    sheet_ws.title = sheet_name

                    # Copy data and formatting from original sheet
                    original_ws = original_wb[sheet_name]

                    # Copy all data with formatting
                    for row in original_ws.iter_rows():
                        for cell in row:
                            # Get the target cell
                            target_cell = sheet_ws.cell(
                                row=cell.row, column=cell.column
                            )

                            # Copy value
                            target_cell.value = cell.value

                            # Copy formatting if cell has style
                            if cell.has_style:
                                # Copy font - create new Font object
                                if cell.font:
                                    target_cell.font = Font(
                                        name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                        color=cell.font.color,
                                    )

                                # Copy border - create new Border object
                                if cell.border:
                                    target_cell.border = Border(
                                        left=cell.border.left,
                                        right=cell.border.right,
                                        top=cell.border.top,
                                        bottom=cell.border.bottom,
                                    )

                                # Copy fill - create new PatternFill object
                                if cell.fill:
                                    target_cell.fill = PatternFill(
                                        fill_type=cell.fill.fill_type,
                                        start_color=cell.fill.start_color,
                                        end_color=cell.fill.end_color,
                                    )

                                # Copy alignment - create new Alignment object
                                if cell.alignment:
                                    target_cell.alignment = Alignment(
                                        horizontal=cell.alignment.horizontal,
                                        vertical=cell.alignment.vertical,
                                        text_rotation=cell.alignment.text_rotation,
                                        wrap_text=cell.alignment.wrap_text,
                                        shrink_to_fit=cell.alignment.shrink_to_fit,
                                        indent=cell.alignment.indent,
                                    )

                                # Copy number format
                                if cell.number_format:
                                    target_cell.number_format = cell.number_format

                                # Copy protection - create new Protection object
                                if cell.protection:
                                    target_cell.protection = Protection(
                                        locked=cell.protection.locked,
                                        hidden=cell.protection.hidden,
                                    )

                    # Copy column dimensions (width)
                    for col in original_ws.column_dimensions:
                        if original_ws.column_dimensions[col].width:
                            sheet_ws.column_dimensions[col].width = (
                                original_ws.column_dimensions[col].width
                            )
                        if original_ws.column_dimensions[col].hidden:
                            sheet_ws.column_dimensions[col].hidden = (
                                original_ws.column_dimensions[col].hidden
                            )

                    # Copy row dimensions (height)
                    for row in original_ws.row_dimensions:
                        if original_ws.row_dimensions[row].height:
                            sheet_ws.row_dimensions[row].height = (
                                original_ws.row_dimensions[row].height
                            )
                        if original_ws.row_dimensions[row].hidden:
                            sheet_ws.row_dimensions[row].hidden = (
                                original_ws.row_dimensions[row].hidden
                            )

                    # Copy merged cells
                    for merged_range in original_ws.merged_cells.ranges:
                        sheet_ws.merge_cells(str(merged_range))

                    # Copy images if any exist
                    if original_ws._images:
                        for img in original_ws._images:
                            sheet_ws.add_image(img)

                    # Save individual XLSX file
                    safe_sheet_name = "".join(
                        c for c in sheet_name if c.isalnum() or c in (" ", "-", "_")
                    ).rstrip()
                    xlsx_file_name = f"{safe_sheet_name}.xlsx"
                    xlsx_file_path = os.path.join(folder_path, xlsx_file_name)

                    sheet_wb.save(xlsx_file_path)
                    print(
                        f"  ✓ Created: {xlsx_file_name} (with images and formatting preserved)"
                    )

                else:
                    # CSV output (original functionality)
                    df = pd.read_excel(xlsm_file_path, sheet_name=sheet_name)

                    # Create CSV file name (replace invalid characters)
                    safe_sheet_name = "".join(
                        c for c in sheet_name if c.isalnum() or c in (" ", "-", "_")
                    ).rstrip()
                    csv_file_name = f"{safe_sheet_name}.csv"
                    csv_file_path = os.path.join(folder_path, csv_file_name)

                    # Save to CSV
                    df.to_csv(csv_file_path, index=False, encoding="utf-8")
                    print(
                        f"  ✓ Created: {csv_file_name} ({len(df)} rows, {len(df.columns)} columns)"
                    )

            except Exception as e:
                print(f"  ✗ Error processing sheet '{sheet_name}': {str(e)}")

        print(f"\nConversion completed! Files saved in: {folder_path}")
        return True

    except Exception as e:
        print(f"Error processing file: {str(e)}")
        return False


def main():
    """Main function to handle command line arguments or interactive input."""
    custom_folder_name = None
    output_format = "xlsx"  # Default to XLSX for image preservation

    if len(sys.argv) > 1:
        # Command line argument provided
        xlsm_file = sys.argv[1]

        # Check for custom folder name argument
        if len(sys.argv) > 2:
            custom_folder_name = sys.argv[2]

        # Check for output format argument
        if len(sys.argv) > 3:
            output_format = sys.argv[3].lower()
            if output_format not in ["xlsx", "csv"]:
                print(
                    "Warning: Invalid output format. Using 'xlsx' for image preservation."
                )
                output_format = "xlsx"
    else:
        # Interactive input
        xlsm_file = input("Enter the path to your XLSM file: ").strip()
        custom_folder_input = input(
            "Enter custom folder name (or press Enter to use file name): "
        ).strip()
        if custom_folder_input:
            custom_folder_name = custom_folder_input

        format_input = (
            input("Enter output format (xlsx/csv, default: xlsx): ").strip().lower()
        )
        if format_input in ["xlsx", "csv"]:
            output_format = format_input

    # Remove quotes if present
    xlsm_file = xlsm_file.strip('"').strip("'")
    if custom_folder_name:
        custom_folder_name = custom_folder_name.strip('"').strip("'")

    # Process the file
    success = xlsm_to_xlsx(xlsm_file, custom_folder_name, output_format)

    if success:
        print("\nScript completed successfully!")
    else:
        print("\nScript failed. Please check the error messages above.")
        sys.exit(1)


if __name__ == "__main__":
    main()
